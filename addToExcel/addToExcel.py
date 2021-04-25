#!/usr/bin/env python3
#! python3
# addToExcel.py - adds data to Excel workbook.

import webbrowser, requests, bs4, sys, os, shutil, pyperclip, openpyxl, logging
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
logging.basicConfig(handlers=[logging.FileHandler('addToExcelLog.txt', 'w', 'utf-8')], level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# List excel files in folder
def list_files():
  for filename in os.listdir('.'):
      if filename.endswith('.xlsx'):
          print(os.path.splitext(filename)[0])
      else:
          continue
        
# Get filename
def get_filename():
    filename = input('\nWhat excel file do you want to add to?\n')
    return filename

print('Here are the excel files in the current folder:')
list_files()


filename = get_filename()
basename = filename

# Make a copy of the original file
if not os.path.exists(os.getcwd() + '/' + basename):
    os.mkdir(basename)
if not os.path.exists(basename + '/' + filename + '.xlsx'):
    shutil.copy(filename + '.xlsx', basename)

wb = openpyxl.load_workbook(filename + '.xlsx')
sheet = wb['Sheet1']

lastCell = get_column_letter(sheet.max_column) + str(sheet.max_row)
nextEmptyRow = str(sheet.max_row + 1)

logging.debug(tuple(sheet['A4':lastCell]))

dataFile = open('contents.txt', encoding='utf-8')
data = dataFile.readlines()
dataFile.close()
logging.debug('Contents: %s\nLength: %d' %(data,len(data)))

rightAlign = Alignment(horizontal='right')

# Fill data
sheet['A' + nextEmptyRow] = data[0].strip() # name
sheet['B' + nextEmptyRow] = data[1].strip() # address
sheet['C' + nextEmptyRow] = data[2].strip() # city
sheet['D' + nextEmptyRow].alignment = rightAlign
zipCode = int(data[4].strip())
sheet['D' + nextEmptyRow] = zipCode # zip code

if len(data) is 6:
  sheet[get_column_letter(sheet.max_column) + nextEmptyRow] = data[5] # phone number
else:
  sheet[get_column_letter(sheet.max_column) + nextEmptyRow] = data[6] # phone number
wb.save(filename + '.xlsx')

# Save excel file to new file
# Figure out the filename this code should use based on what files already exist.
number = 1
while True:
    filename = basename + '_' + str(number) + '.xlsx'
    logging.debug('Filename: %s' %filename)
    if not os.path.exists(basename + '/' + filename):
      break
    number = number + 1
    logging.debug('number: %d filename: %s' %(number, filename))
wb.save(filename)
shutil.move(filename, basename)
wb.close()
print('File saved to %s' %filename)
input('Press ENTER to exit...')
